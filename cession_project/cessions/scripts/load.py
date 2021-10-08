import datetime

from django.shortcuts import get_object_or_404

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.read_only import EmptyCell


from cessions.models import Territory, Publishing_company, Author, Illustrator, Topic, Collection, Book, Agency, Agent, Royalties, Cession, Publisher_contact

def run():

    wb = load_workbook(filename = 'E:\OPS\cession\cession_project\cessions\scripts\Tableau_de_bord_cessions_complet.xlsx')
    sheet_ranges_suivi = wb['TdB suivi']
    sheet_ranges_cessions = wb['cessions']
    
    Book.objects.all().delete()
    Collection.objects.all().delete()
    Publishing_company.objects.all().delete()
    Territory.objects.all().delete()
    Agency.objects.all().delete()
    Publisher_contact.objects.all().delete()
    Agent.objects.all().delete()
    Cession.objects.all().delete()
    

    # Sheet : Cessions

    for r in range(2, 299):
        type = sheet_ranges_cessions.cell(row = r, column = 1)
        str_published_year = sheet_ranges_cessions.cell(row = r, column = 2)
        if str_published_year.value is not None:
            date_published_year = datetime.datetime.strptime(str_published_year.value, "%Y").date()
        title = sheet_ranges_cessions.cell(row = r, column = 3)

        if type.value is not None:
            if type.value == "A":
                collection = "Album"
            if type.value == "R":
                collection = "Roman"
            if type.value == "L":
                collection = "Loupio"
            obj_collection, created = Collection.objects.get_or_create(name = collection)
        if title.value is not None:
            obj_book, created = Book.objects.get_or_create(title = title.value, published_at = date_published_year, collection = obj_collection)

    # Sheet : TdB suivi

    for r in range(2, 1465):
        occasion = sheet_ranges_suivi.cell(row = r, column = 1)
        comment = sheet_ranges_suivi.cell(row = r, column = 2)
        book_title = sheet_ranges_suivi.cell(row = r, column = 3)
        territory_name = sheet_ranges_suivi.cell(row = r, column = 4)
        publishing_company_name = sheet_ranges_suivi.cell(row = r, column = 5)
        agency_name = sheet_ranges_suivi.cell(row = r, column = 6)
        contact_name = sheet_ranges_suivi.cell(row = r, column = 7)
        contact_email = sheet_ranges_suivi.cell(row = r, column = 8)

        pdf = sheet_ranges_suivi.cell(row = r, column = 9)
        contract_sent_at = sheet_ranges_suivi.cell(row = r, column = 11)
        contract_signed_at = sheet_ranges_suivi.cell(row = r, column = 12)
        bill_sent_at = sheet_ranges_suivi.cell(row = r, column = 13)

        book_sent_at = sheet_ranges_suivi.cell(row = r, column = 14)
        files_sent_at = sheet_ranges_suivi.cell(row = r, column = 16)

        model_received_at = sheet_ranges_suivi.cell(row = r, column = 17)

        try:
            
            pdf_send_at = datetime.datetime.strptime(pdf.value, "%d/%m/%Y").date()
        except:
            pdf_send_at = None

        try:
            contract_sent_at = datetime.datetime.strptime(contract_sent_at.value, "%d/%m/%Y").date()
        except:
            contract_sent_at = None

        try:
            contract_signed_at = datetime.datetime.strptime(contract_signed_at.value, "%d/%m/%Y").date()
        except:
            contract_signed_at = None

        try:
            bill_sent_at = datetime.datetime.strptime(bill_sent_at.value, "%d/%m/%Y").date()
        except:
            bill_sent_at = None

        try:
            book_sent_at = datetime.datetime.strptime(book_sent_at.value, "%d/%m/%Y").date()
        except:
            book_sent_at = None

        try:
            files_sent_at = datetime.datetime.strptime(files_sent_at.value, "%d/%m/%Y").date()
        except:
            files_sent_at = None

        try:
            model_received_at = datetime.datetime.strptime(model_received_at.value, "%d/%m/%Y").date()
        except:
            model_received_at = None

        if contact_name.value is not None:
            contact_first_name = (contact_name.value + " ").split(" ")[0]
            contact_last_name = (contact_name.value + " ").split(" ")[1:-1]
            contact_last_name = " ".join(contact_last_name)

        a_created = False
        pc_created = False
        agent_created = False
        publisher_created = False

        if territory_name.value is not None:
            t, created = Territory.objects.get_or_create(name=territory_name.value)
            if publishing_company_name.value is not None:
                pc, pc_created = Publishing_company.objects.get_or_create(name=publishing_company_name.value)
                if pc_created == True:
                    pc.territories.add(t)
                    pc.save()
                    publisher_object, publisher_created = Publisher_contact.objects.get_or_create(first_name=contact_first_name, last_name=contact_last_name, email=contact_email.value, publishing_company=pc)
            if agency_name.value is not None:
                a, a_created = Agency.objects.get_or_create(name=agency_name.value)
                if a_created == True:
                    a.territories.add(t)
                    a.save()
                    agent_object, agent_created = Agent.objects.get_or_create(first_name=contact_first_name, last_name=contact_last_name, email=contact_email.value, agency=a)
        
        if book_title.value is not None:
            str_book_title = book_title.value
            
            if len(str_book_title) > 0:
                obj_book, book_created = Book.objects.get_or_create(title = str_book_title)

            if  a_created == True and pc_created == False :
                try:
                    obj_cession, created = Cession.objects.get_or_create(book = obj_book, agency = a, created_at = pdf_send_at, pdf_sent_at = pdf_send_at, contract_sent_at = contract_sent_at, contract_signed_at = contract_signed_at, bill_sent_at = bill_sent_at, book_sent_at = book_sent_at, files_sent_at = files_sent_at, model_received_at = model_received_at, language = territory_name.value, occasion = occasion.value, comment = comment.value )
                except:
                    print("cession {} not creted".join(obj_book.title))

            elif a_created == False and pc_created == True :
                try:
                    obj_cession, created = Cession.objects.get_or_create(book = obj_book, publishing_company = pc, created_at = pdf_send_at, pdf_sent_at = pdf_send_at, contract_sent_at = contract_sent_at, contract_signed_at = contract_signed_at, bill_sent_at = bill_sent_at, book_sent_at = book_sent_at, files_sent_at = files_sent_at, model_received_at = model_received_at, language = territory_name.value, occasion = occasion.value, comment = comment.value )
                except:
                    print("cession {} not creted".join(obj_book.title))
            elif a_created == False and pc_created == True :
                try:
                    obj_cession, created = Cession.objects.get_or_create(book = obj_book, agency = a, publishing_company = pc, created_at = pdf_send_at, pdf_sent_at = pdf_send_at, contract_sent_at = contract_sent_at, contract_signed_at = contract_signed_at, bill_sent_at = bill_sent_at, book_sent_at = book_sent_at, files_sent_at = files_sent_at, model_received_at = model_received_at, language = territory_name.value, occasion = occasion.value, comment = comment.value )
                except:
                    print("cession {} not creted".join(obj_book.title))                
            else:
                try:
                    obj_cession, created = Cession.objects.get_or_create(book = obj_book, created_at = pdf_send_at, pdf_sent_at = pdf_send_at, contract_sent_at = contract_sent_at, contract_signed_at = contract_signed_at, bill_sent_at = bill_sent_at, book_sent_at = book_sent_at, files_sent_at = files_sent_at, model_received_at = model_received_at, language = territory_name.value, occasion = occasion.value, comment = comment.value )
                except:
                    print("cession {} not creted".join(obj_book.title))         
            if created == True:
                    obj_cession.territories.add(t)
                    if agent_created == True:
                        obj_cession.agent = agent_object
                    if publisher_created == True:
                        obj_cession.publisher_contact = publisher_object
                    obj_cession.save()